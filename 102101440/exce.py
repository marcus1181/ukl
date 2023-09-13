import requests
from bs4 import BeautifulSoup
import re
import jieba
import wordcloud
from collections import Counter
import imageio
from openpyxl import Workbook,load_workbook
from PIL import Image
import numpy as np
from openpyxl import Workbook,load_workbook
exce=Workbook()
sheet=exce.active
print(sheet.title)
sheet.title="b站弹幕搜集"
sheet["A1"]="编号"
sheet["B1"]="弹幕"
sheet["C1"]="出现次数"
exce.save("日本和污染水排海弹幕爬取.xlsx")
i=1
f = open('日本核污染水排海弹幕.txt','r',encoding= 'utf-8')
a = f.read()
list = a.split("\n")
for i in range(0,1000) :
    sheet.append([i+1,list[i],list.count(list[i])])
print("=============生成日本和污染水排海弹幕爬取.xlsx文件===========")
exce.save("日本和污染水排海弹幕爬取.xlsx")